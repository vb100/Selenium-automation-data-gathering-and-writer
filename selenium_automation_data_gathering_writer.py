# Import libraries
import requests, re, os
import pandas as pd
from bs4 import BeautifulSoup
import os

print("Starting sheet: Inflation rates")

""" Prepare Home directory : start """
os.chdir("C:\\Users\\Vytautas.Bielinskas\\Desktop\\PythonWorking\\Python\\")

""" --------------------------------------------------------------------------------------"""
""" Read data from source : start """
def Genesis_Table01():
    
    # Import libraries
    import requests, re, os
    import pandas as pd
    from bs4 import BeautifulSoup
    import os
    
    """ Prepare Home directory : start """
    os.chdir("C:\\Users\\Vytautas.Bielinskas\\Desktop\\PythonWorking\\Python\\")
    
    from selenium import webdriver
    options = webdriver.chrome.options.Options()
    options.add_argument("--disable-extensions")
        
    chrome_path = r"C:\Users\Vytautas.Bielinskas\Desktop\Python\JSscrapping\chromedriver.exe"
    driver = webdriver.Chrome(chrome_path)
    
    """ -----------------------------------------------------------------------------------"""
    """ Working with Table : start """
    def parseTable(table):
        table = table.split("<tr>")
        
        row = {}
        l = []
        
        for item in range(1, len(table), 1):
            string = table[item]
            
            if '<acronym title="numerical value unknown or not to be disclosed">'.upper() in string.upper():
                cutting = '<acronym title="numerical value unknown or not to be disclosed">'
                string = string.split(cutting)[0] + string.split(cutting)[1]  
                
            row_string = string.split('</th>')
            
            year = row_string[0].split('">')[len(row_string[0].split('">'))-1]
            row["Year"] = year
            month = row_string[1].split('">')[len(row_string[1].split('">'))-1]
            row["Months"] = month
            
            list_of_last_part = row_string[2].split("</td>")
            for i in range(0, len(list_of_last_part)-1, 1):
                value = list_of_last_part[i].split('">')[1]
                if "</acronym>".upper() in value.upper():
                    value = value.split("</acronym>")[0]
                        
                if i == 0:
                    if len(value) > 5:
                        value = "-"
                    row["CPI"] = value
                elif i == 1:
                    if value == "." or len(value) > 5:
                        value = "-"                
                    row["CpyM"] = value
                elif i == 2:
                    if len(value) > 5:
                        value = "-"
                    row["CpM"] = value
                    
            l.append(dict(row))
            
        df = pd.DataFrame(l)
        df = df[["Year", "Months", "CPI", "CpyM","CpM"]]
    
        # Fill the date column
        start_date = []
        for i in range(0, len(df), 1):
            if len(df.iat[i,0]) > 2:
                start_date.append(i)
        
        for i in range(0, len(start_date)-1, 1):
            year = df.iat[start_date[i], 0]
            for j in range(start_date[i], start_date[i+1], 1):
                df.iat[j,0] = year
        
        print(start_date[len(start_date)-1], len(df))
        
        for j in range(start_date[len(start_date)-1], len(df), 1):
            df.iat[j,0] = df.iat[start_date[len(start_date)-1], 0]
        
        last_date = str(df.iat[len(df)-1, 0]) + df.iat[len(df)-1, 1]
        
        print("The last date is:", last_date)        
        
        return df
    """ Working with Table : end """
    """ -----------------------------------------------------------------------------------"""
    """ -----------------------------------------------------------------------------------"""
    """ Writing to Excel file : start """
    def writingToExcel(df):
        import xlwings as xw
        import openpyxl, os, datetime
        from openpyxl.utils import get_column_letter
        from openpyxl import Workbook
        from pandas import ExcelWriter as ewriter
        from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Fill
        
        """ Define some styles : start """
        
        color_value = 5
        
        #ft = Font(name = "Arial",
        #          size = 8,
        #          bold = False,
        #          italic = False,
        #          color = "3385ff")
        
        #rightAligment = Alignment(horizontal = "right")
        """ Define some styles : end """
    
        os.chdir("C:\\Users\\Vytautas.Bielinskas\\Desktop\\PythonWorking\\Python\\")
        filename = "Germany macro - Aug 2017.xlsx"  
        
        wb = xw.Book(filename)
        
        print(df)
        print("Enter key to continue")
        
        for i in range(0, len(df), 1):
            for j in range(0, 5, 1):
                wb.sheets['Inflation rates'].range(i+14, j+1).value = df.iat[i, j]
                print(i, ",", j, "-->", df.iat[i, j])
                if j == 2 or j == 3 or j == 4:
                    wb.sheets['Inflation rates'].range(i+14, j+1).number_format = "0.0"
                print("keeping...")

        row_last = 14
        print("Just checking...")
        print(str(wb.sheets['Inflation rates'].range(row_last, 11).value))
        while "Q" in (str(wb.sheets['Inflation rates'].range(row_last, 11).value)):
            row_last = row_last + 1
        
        print("row_i", row_last)
        
        if "2017" in table_df.iat[len(df)-1, 0]:
            if "December".upper() in df.iat[len(df)-1, 1].upper():
                #sheet.cell(row = row_last, column = 11).value = "2017 Q4"
                wb.sheets['Inflation rates'].range(row_last, 11).value = "2017 Q4"
                wb.sheets['Inflation rates'].range(row_last, 11).api.Font.ColorIndex = color_value
        
        # Defining F column
        print("Defining column F:", 20, len(table_df)+13, 1)
        for row_i in range(20, len(table_df)+13, 1):          
            print("Lenght:", len(str(wb.sheets['Inflation rates'].range(row_i, 6).value).replace(" ", "").split(":")[0].replace("-", "")), " | ", str(wb.sheets['Inflation rates'].range(row_i, 6).value).replace(" ", "").split(":")[0].replace(" ", ""))
            if len(str(wb.sheets['Inflation rates'].range(row_i, 6).value).replace(" ", "").split(":")[0].replace(" ", "")) < 11:
                year = wb.sheets['Inflation rates'].range(row_i, 1).value
                
                # Getting month
                print("MONTH V:", wb.sheets['Inflation rates'].range(row_i, 2).value, "Year:", year, type(year))
                month_v = str(wb.sheets['Inflation rates'].range(row_i, 2).value)
                if "Jan".upper() in month_v.upper():
                    month =  "01"
                elif "Feb".upper() in month_v.upper():
                    month = "02"
                elif "March".upper() in month_v.upper():
                    month = "03"
                elif "Apri".upper() in month_v.upper():
                    month = "04"
                elif "May".upper() in month_v.upper():
                    month = "05"
                elif "Jun".upper() in month_v.upper():
                    month = "06"
                elif "Jul".upper() in month_v.upper():
                    month = "07"
                elif "Augu".upper() in month_v.upper():
                    month = "08"
                elif "Sept".upper() in month_v.upper():
                    month = "09"
                elif "Octob".upper() in month_v.upper():
                    month = "10"
                elif "Nove".upper() in month_v.upper():
                    month = "11"
                elif "Decem".upper() in month_v.upper():
                    month = 12
                
                print("Year:", year, "-->", month + "/01/" + str(year).replace(".0", ""))
                wb.sheets['Inflation rates'].range(row_i, 6).value = str(month + "/01/" + str(year)).replace(".0", "")
                wb.sheets['Inflation rates'].range(row_i, 6).api.Font.ColorIndex = color_value
                
        """ Writing L, N, O column : start """
        for row_i in range(14, row_last, 1):
            formula_L = '=AVERAGEIF($H$14' + ':$H$' + str(len(df) + 12) + ',K' + str(row_i) + ',$I$14:$I$' + str(len(df) + 12) + ')'
            formula_N = '=RIGHT(K' + str(row_i) + ',2)&" "&LEFT(K' + str(row_i) + ',4)'
            formula_O = '=L' + str(row_i)
            formula_P = '=VLOOKUP(N' + str(row_i) + ',$BD$' + str(18) + ':$BE$' + str(row_last) + ',2,FALSE)' #row_last pakeisti i dinamiska
    
            wb.sheets['Inflation rates'].range(row_i, 12).value = formula_L
            wb.sheets['Inflation rates'].range(row_i, 12).api.Font.ColorIndex = color_value
            
            wb.sheets['Inflation rates'].range(row_i, 14).value = formula_N
            wb.sheets['Inflation rates'].range(row_i, 14).api.Font.ColorIndex = color_value
            
            wb.sheets['Inflation rates'].range(row_i, 15).value = formula_O
            wb.sheets['Inflation rates'].range(row_i, 15).api.Font.ColorIndex = color_value
            
            wb.sheets['Inflation rates'].range(row_i, 16).value = formula_P
            wb.sheets['Inflation rates'].range(row_i, 16).api.Font.ColorIndex = color_value
        """ Writing L column : end """
                
        for i in range(1, len(df), 1):
            
            formula_1 = '=A' + str(i + 13) + '&' + '" "&"Q"&G' + str(i + 13)
            wb.sheets['Inflation rates'].range(i + 13, 8).value = formula_1
            wb.sheets['Inflation rates'].range(i + 13, 8).api.Font.ColorIndex = color_value
            
            formula_2 = '=ROUNDUP(MONTH(F' + str(i + 13) + ')/3,0)'
            wb.sheets['Inflation rates'].range(i + 13, 7).value = formula_2
            wb.sheets['Inflation rates'].range(i + 13, 7).api.Font.ColorIndex = color_value
            
            # CHECKING D AND I COLUMNS              
            if wb.sheets['Inflation rates'].range(i + 13, 4).value == ".":
                wb.sheets['Inflation rates'].range(i + 13, 9).value = ""
                wb.sheets['Inflation rates'].range(i + 13, 9).api.Font.ColorIndex = color_value
            elif wb.sheets['Inflation rates'].range(i + 13, 4).value == "":
                wb.sheets['Inflation rates'].range(i + 13, 9).value = float(0)
                wb.sheets['Inflation rates'].range(i + 13, 9).api.Font.ColorIndex = color_value
            else:
                print("See here:", i+14, ":", wb.sheets['Inflation rates'].range(i + 14, 4).value, type(wb.sheets['Inflation rates'].range(i + 14, 4).value))
                
                if "-" in str(wb.sheets['Inflation rates'].range(i + 13, 4).value) and len(str(wb.sheets['Inflation rates'].range(i + 13, 4).value)) == 1:
                    wb.sheets['Inflation rates'].range(i + 13, 9).value = 0
                elif "-" in str(wb.sheets['Inflation rates'].range(i + 13, 4).value) and len(str(wb.sheets['Inflation rates'].range(i + 13, 4).value)) > 1:
                    wb.sheets['Inflation rates'].range(i + 13, 9).value = float(str(wb.sheets['Inflation rates'].range(i + 13, 4).value))
                else:
                    wb.sheets['Inflation rates'].range(i + 13, 9).value = float(str(wb.sheets['Inflation rates'].range(i + 13, 4).value).replace("+", ""))
                wb.sheets["Inflation rates"].range(i + 13, 9).api.Font.ColorIndex = color_value
        
        #wb.save("reuslt-FINISHED.xlsx")
        wb.save()
        print("Inflation rates 1: writing is done.")  
        
        return None
    """ Writing to Excel file : end """
    """ -----------------------------------------------------------------------------------"""
    
    driver.get("https://www-genesis.destatis.de")
    
    """ MAIN SEACRH PAGE """
    input_field = driver.find_element_by_xpath('//*[@id="suchanweisung"]')
    search_button = driver.find_element_by_xpath('//*[@id="btnSucheAbsenden"]')
    table_code = "61111-0002"
    
    print(driver.current_url)
    
    input_field.clear()
    input_field.send_keys(table_code)
    search_button.click()
    
    print(driver.current_url)
    
    """ RESULT PAGE """
    link_to_result = driver.find_element_by_xpath('//*[@id="wrapperContent"]/form/table/tbody/tr[2]/td[2]/a')
    link_to_result.click()
    
    print(driver.current_url)
    
    """ DATASET PAGE """
    set_years = driver.find_element_by_xpath('//*[@id="wrapperContent"]/form[1]/table/tbody/tr[8]/td[4]/div/input')
    set_years.click()
    
    print(driver.current_url)
    
    """ SET YEARS PAGE """
    all_period = driver.find_element_by_xpath('//*[@id="za_1"]')
    accept_button = driver.find_element_by_xpath('//*[@id="wrapperContent"]/form/div/div[3]/div[1]/input')
    all_period.click()
    accept_button.click()
    print('Accept - clicked')
    
    print(driver.current_url)
    
    """ Come back to DATASET PAGE """
    value_retr_button = driver.find_element_by_xpath('//*[@id="wrapperContent"]/form[1]/div[7]/input[2]')
    value_retr_button.click()

    """ Get the table! """
    table = driver.find_element_by_xpath('//*[@id="wrapperContent"]/div[3]/form/table/tbody')
    table_s = table.get_attribute("innerHTML")
    
    #print(table_s)
    driver.close()
    
    table_df = parseTable(table_s)
    writingToExcel(table_df)
    
    #print(table.text)
    
    return None
""" Read data from source : end """ 
""" --------------------------------------------------------------------------------------"""
""" --------------------------------------------------------------------------------------"""
def Genesis_Table02():
    
    # Import libraries
    import requests, re, os
    import pandas as pd
    from bs4 import BeautifulSoup
    import os
    
    from selenium import webdriver
    options = webdriver.chrome.options.Options()
    options.add_argument("--disable-extensions")
        
    chrome_path = r"C:\Users\Vytautas.Bielinskas\Desktop\Python\JSscrapping\chromedriver.exe"
    driver = webdriver.Chrome(chrome_path)
    
    """ Writing to Excel file : start """
    def WritingToExcel(df):
        
        """ get Column Letter : start """
        def getColumnLetter(number):
            #print("Getting letter from:", number)
            if number == 1:
                Letter = "A"
            elif number == 2:
                Letter = "B"
            elif number ==3:
                Letter = "C"
            elif number ==4:
                Letter = "D"
            elif number ==5:
                Letter = "E"
            elif number ==6:
                Letter = "F"
            elif number ==7:
                Letter = "G"
            elif number ==8:
                Letter = "H"
            elif number ==9:
                Letter = "I"
            elif number ==10:
                Letter = "J"
            elif number ==11:
                Letter = "K"
            elif number ==12:
                Letter = "L"
            elif number ==13:
                Letter = "M"
            elif number ==14:
                Letter = "N"
            elif number ==15:
                Letter = "O"
            elif number ==16:
                Letter = "P"
            elif number ==17:
                Letter = "Q"
            elif number ==18:
                Letter = "R"
            elif number ==19:
                Letter = "S"
            elif number ==20:
                Letter = "T"
            elif number ==21:
                Letter = "U"
            elif number ==22:
                Letter = "V"
            elif number ==23:
                Letter = "W"
            elif number ==24:
                Letter = "X"
            elif number ==25:
                Letter = "Y"
            elif number ==26:
                Letter = "Z"
            elif number ==27:
                Letter = "AA"
            elif number ==28:
                Letter = "AB"
            elif number ==29:
                Letter = "AC"
            elif number ==30:
                Letter = "AD"
            elif number ==31:
                Letter = "AE"
            elif number ==32:
                Letter = "AF"
            elif number ==33:
                Letter = "AG"
            elif number ==34:
                Letter = "AH"
            elif number ==35:
                Letter = "AI"
            elif number ==36:
                Letter = "AJ"
            elif number ==37:
                Letter = "AK"
            elif number ==38:
                Letter = "AL"
            elif number ==39:
                Letter = "AM"
            elif number ==40:
                Letter = "AN"
            elif number ==41:
                Letter = "AO"
            elif number ==42:
                Letter = "AP"
            elif number ==43:
                Letter = "AQ"
            elif number ==44:
                Letter = "AR"
            elif number ==45:
                Letter = "AS"
            elif number ==46:
                Letter = "AT"
            elif number ==47:
                Letter = "AU"
            elif number ==48:
                Letter = "AV"
            elif number ==49:
                Letter = "AW"
            elif number ==50:
                Letter = "AX"
            elif number ==51:
                Letter = "AY"
            elif number ==52:
                Letter = "AZ"
            elif number ==53:
                Letter = "BA"
            elif number ==54:
                Letter = "BB"
            elif number ==55:
                Letter = "BC"
            elif number ==56:
                Letter = "BD"
            elif number ==57:
                Letter = "BE"
            elif number ==58:
                Letter = "BF"
            elif number ==59:
                Letter = "BG"
            return Letter
        """ get Column Letter : end """
        
        import xlwings as xw
        import openpyxl, os, datetime
        from openpyxl.utils import get_column_letter
        from openpyxl import Workbook
        from pandas import ExcelWriter as ewriter
        from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Fill
        
        print("Writing the second part of Excel file")
        
        """ Define some styles : start """
        #ft = Font(name = "Arial",
        #          size = 8,
        #          bold = False,
        #          italic = False,
        #          color = "3385ff")
        
        #rightAligment = Alignment(horizontal = "right")
        
        color_value = 3
        """ Define some styles : end """
        
        os.chdir("C:\\Users\\Vytautas.Bielinskas\\Desktop\\PythonWorking\\Python\\04 Germany Macro")
        filename = "Germany macro - Aug 2017.xlsx"  
        
        #wb = openpyxl.load_workbook(filename)
        #wb.get_sheet_names()
        wb = xw.Book(filename)
        print("|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||")
        print("<<<Excel file is opened now!>>>")
        
        # Fill the Table!       
        for row_i in range(14, 41, 1):                            # Till 2017 Years!
            for col in range(23, 35, 1):
                #sheet.cell(row = row_i, column = col).value = float(df.iat[row_i - 14, col - 23])
                #sheet.cell(row = row_i, column = col).font = ft
                wb.sheets['Inflation rates'].range(row_i, col).value = float(df.iat[row_i - 14, col - 23])
                wb.sheets['Inflation rates'].range(row_i, col).api.Font.ColorIndex = color_value
            #print(row_i, "row is done.")
                
                
        col = 37
        row_i = 15
               
        print("Next step.")
        for row_i in range(15, 41, 1):
            print(row_i, "row proceesing...")
            for col in range(37, 49, 1):
                print(col, "column proceesing...")
                letter = getColumnLetter(col-14)
                formula = '=(' + letter + str(row_i) + '-' + letter + str(row_i-1) + ')/' + letter + str(row_i-1) + '* 100'
                print(formula)
                wb.sheets['Inflation rates'].range(row_i, col).value = formula
                print("formula is written: done")
                try:
                    checking_value = wb.sheets['Infliation rates'].range(row_i, col).value
                    if checking_value == "-100":
                        print("If conditions is TRUE")
                        wb.sheets['Infliation rates'].range(row_i, col).value = ""
                        print("None value is written.")
                except:
                    print("Everything it's OK")
                print("Color is applying...")
                wb.sheets['Inflation rates'].range(row_i, col).api.Font.ColorIndex = color_value
                print("Color is applied.")       
        
        """ Writing AX column : start """
        
        print("Step 3")
        for row_i in range(14, 41, 1):
            wb.sheets['Inflation rates'].range(row_i, 50).value = '=S' + str(row_i)  # Till 2017 years
            wb.sheets['Inflation rates'].range(row_i, 50).api.Font.ColorIndex = color_value    
            print(row_i, "row is done.")
            
        """ Writing AY column Q1: start """
            
        print("Step 4")
        for row_i in range(15, 41, 1):
            wb.sheets['Inflation rates'].range(row_i, 51).value = '=AVERAGE(AK' + str(row_i) + ':AM' + str(row_i) + ')'
            wb.sheets['Inflation rates'].range(row_i, 51).api.Font.ColorIndex = color_value
            print(row_i, "row is done.")
            
        """ Writing AZ column Q2: start """
            
        print("Step 5")
        for row_i in range(15, 41, 1):
            wb.sheets['Inflation rates'].range(row_i, 52).value = '=AVERAGE(AN' + str(row_i) + ':AP' + str(row_i) + ')'
            wb.sheets['Inflation rates'].range(row_i, 52).api.Font.ColorIndex = color_value  
            print(row_i, "row is done")
            
        """ Writing BA column Q3: start """
            
        print("Step 6")
        for row_i in range(15, 41, 1):
            wb.sheets['Inflation rates'].range(row_i, 53).value = '=AVERAGE(AQ' + str(row_i) + ':AS' + str(row_i) + ')'
            wb.sheets['Inflation rates'].range(row_i, 53).api.Font.ColorIndex = color_value 
            print(row_i, "row is done.")
            
        """ Writing BA column Q4: start """
            
        print("Step 7")
        for row_i in range(15, 41, 1):
            wb.sheets['Inflation rates'].range(row_i, 54).value = '=AVERAGE(AT' + str(row_i) + ':AV' + str(row_i) + ')'
            wb.sheets['Inflation rates'].range(row_i, 54).api.Font.ColorIndex = color_value
            print(row_i, "row is done.")
                
        wb.save()
        print("File is saved.")
        #wb.save(filename + "-FINISHED.xlsx")
        print("Inflation rates 2: writing is done.")        
        
        return None
    """ Writing to Excel file : end """

    driver.get("https://www-genesis.destatis.de")
    
    """ MAIN SEARCH PAGE"""
    input_field = driver.find_element_by_xpath('//*[@id="suchanweisung"]')
    search_button = driver.find_element_by_xpath('//*[@id="btnSucheAbsenden"]')
    table_code = "61131-0002"
    
    input_field.clear()
    input_field.send_keys(table_code)
    search_button.click()
    
    """ RESULT PAGE """
    link_to_result = driver.find_element_by_xpath('//*[@id="wrapperContent"]/form/table/tbody/tr[2]/td[2]/a')
    link_to_result.click()
    
    print(driver.current_url)
    
    """ SELECT YEARS BUTTON """
    year_button = driver.find_element_by_xpath('//*[@id="wrapperContent"]/form[1]/table/tbody/tr[5]/td[4]/div/input')
    year_button.click()
    
    print(driver.current_url)
    
    """ SELECT YEARS """
    select_years = driver.find_element_by_xpath('//*[@id="za_1"]')
    accept_years = driver.find_element_by_xpath('//*[@id="wrapperContent"]/form/div/div[3]/div[1]/input')
    select_years.click()
    accept_years.click()
    
    print(driver.current_url)
    
    """ SELECT VAT """
    select_vat = driver.find_element_by_xpath('//*[@id="wrapperContent"]/form[1]/table/tbody/tr[7]/td[4]/div/input')
    select_vat.click()
    
    print(driver.current_url)
    
    select_vat_radio = driver.find_element_by_xpath('//*[@id="newselected1"]')
    select_vat_radio.click()
    
    accept_button = driver.find_element_by_xpath('//*[@id="wrapperContent"]/form/div[6]/div[1]/input')
    accept_button.click()
    
    """ SELECT WZ2008 --> WZ08-47-01 """
    selected_item = driver.find_element_by_xpath('//*[@id="wrapperContent"]/form[1]/table/tbody/tr[8]/td[4]/div/input')
    selected_item.click()
    print('Select item - clicked')
    
    wz08_47_01 = driver.find_element_by_xpath('//*[@id="newselected1"]')
    wz08_47_01.click()
    print('wz08_47_01 - clicked')
    
    accept_button = driver.find_element_by_xpath('//*[@id="wrapperContent"]/form/div[8]/div[1]/input')
    accept_button.click()
    print('Accept - clicked')
    
    """ VALUE RETRIEVAL """
    accept_button = driver.find_element_by_xpath('//*[@id="wrapperContent"]/form/div[6]/div[1]/input')
    accept_button.click()
    print(driver.current_url)
    print('Accept - clicked')
    
    view_table = driver.find_element_by_xpath('//*[@id="wrapperContent"]/form[1]/div[7]/input[2]')
    view_table.click()
    print('View values - clicked')
    print(driver.current_url)
    
    """ GET THE TABLE!"""
    table = driver.find_element_by_xpath('//*[@id="wrapperContent"]/div[3]/form/table/tbody')
    table_s = table.get_attribute("innerHTML")
    
    """ Reorganizing Data! """
    table_l = table_s.split("</td>")
    listing = []
    
    row = 0
    while row < len(table_l)-1:
        value = table_l[row].split('">')[len(table_l[row].split('">'))-1]
        if "acronym" in value or "</tr>" in value:
            value = "0.0"
        print((row % 12) + 1, ":", value)
        #record[(row % 12) + 1] = value
        listing.append(value)
        row = row + 1 
        
    record = {}
    book = []
    index = 0
    for i in range(0, round(len(listing)/12), 1):
        print(i,"--:")
        for j in range(0, 12, 1):
            record[j] = listing[index]
            print(index, ":", listing[index])
            index = index + 1
        book.append(dict(record))
    db = pd.DataFrame(book)
    
    db["Year"] = ""
    for i in range(0, len(db), 1):
        db.iat[i, 12] = 1991 + i
        
    driver.close()
        
    WritingToExcel(db)
    
    return None
""" --------------------------------------------------------------------------------------"""

df_01 = Genesis_Table01()
df_02 = Genesis_Table02()